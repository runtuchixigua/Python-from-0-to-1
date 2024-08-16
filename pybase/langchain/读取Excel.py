from langchain_community.llms import Tongyi
from openpyxl import load_workbook

llm = Tongyi(
    model='qwen-max',
    dashscope_api_key='sk-188e803b54354551bdae6b708daa20da'
)

wb = load_workbook('file_path_feedback.xlsx')
ws = wb.active
for row in ws.iter_rows(min_row=1, values_only=True):
    text = row[0]
    result = llm.invoke(f"""
    你需要对用户的反馈进行原因分类。
    分类包括:价格过高、售后支持不足、产品使用体验不佳、其他。
    回答格式为:分类结果:xx。
    用户的问题是:{text}
    """)
    print(result)
    print('----------------------------------')
